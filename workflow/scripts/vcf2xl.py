from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import argparse, re
import sys

parser = argparse.ArgumentParser()
parser.add_argument("vcf")
parser.add_argument("--fields", nargs="*", default=[])
parser.add_argument("--fieldnames", nargs="*", default=[])
parser.add_argument("--genotypes", nargs="*", default=[])
parser.add_argument("-out")
args = parser.parse_args()
seperators = [",", "|", ";"]

#output_infos = ["ANN[*][3]", "gnomad_AF"]
filename = args.vcf
output_infos = args.fields
genotype_infos = args.genotypes
fieldnames = args.fieldnames
use_infos = []
use_genotypes = []

assert(len(fieldnames) == len(output_infos) + len(genotype_infos))

for x in output_infos:
    name = re.match(r"\w*", x).group() 
    matches = re.findall(r"\[[0-9]+\]|\[\*\]", x)
    indices = []
    for m in matches:
        value = m[1:-1]
        if value.isnumeric():
            value = int(value)
        indices.append(value)
    use_infos.append((name, indices))


for x in genotype_infos:
    name = re.match(r"\w*", x).group()
    rest = x[len(name):]
    if len(rest) > 0:
        assert(rest[0] == "[" and rest[-1] == "]")
        rest = rest[1:-1]
        use_genotypes.append((name, rest.split(",")))
    else:
        use_genotypes.append((name, []))


def parse(indices, seperators, s):
    if len(indices) == 0:
        return s
    current_index = indices[0]
    current_sep = seperators[0]
    split = s.split(current_sep)
    if current_index == "*":
        return [parse(indices[1:], seperators[1:], s) for s in split]
    else:
        return split[current_index]


def get_gt_sum(s, gt):
    ret = []
    for g in gt:
        rg = re.escape(g)
        m = re.search("[0-9]*" + rg ,s)
        if m:
            ret.append(int(m.group(0)[:-len(g)]))
    return sum(ret)

# def parse_genotype(obs):
#     #8N-4V-4N+2V+

def convfloat(s):
    if s == "":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return s

wb = Workbook()
ws = wb.active
ws.title = "denovo"

if filename == "/dev/stdin" or filename == "-":
    f = sys.stdin
else:
    f = open(filename, 'r')

outline = ["id", "chrom", "pos", "ref", "alt", "qual"] + fieldnames[:len(output_infos)]

row = 2
n_id = 0
for line in f:
    if line.startswith("##"):
        continue
    if line.startswith("#"):
        samples = line.strip().split("\t")[9:]
        for prefix in fieldnames[len(output_infos):]:
            for s in samples:
                outline.append(prefix + s)
        ws.append(outline)
        continue
    split = line.strip().split("\t")
    CHROM, POS, ID, REF, ALT, QUAL, FILTER, INFO, FORMAT = split[0:9]
    GENOTYPES = split[9:]
    ws.append([n_id, CHROM, POS, REF, ALT, QUAL])
    col=7
    is_single_value = [1,2,3,4,5,6]

    info = {}
    for x in INFO.split(";"):
        x_split = x.split("=", 1)
        x0 = x_split[0]
        if len(x_split) == 2:
            x1 = x_split[1]
        else:
            x1 = True
        info[x0] = x1

    next_row = row + 1
    for variable, indices in use_infos:
        values = parse(indices, seperators, info.get(variable, ""))
        if isinstance(values, list):
            for i, value in enumerate(values):
                ws.cell(column=col, row=row + i, value=value)
                next_row = max(next_row, row + 1 + i)
        else:
            ws.cell(column=col, row=row, value=convfloat(values))
            is_single_value.append(col)
        col+=1
        #outline.append(parse(indices, seperators, value))

    d = {f:i for i,f in enumerate(FORMAT.split(":"))}
    for name, inside in use_genotypes:
        name_index = d[name]
        for g in GENOTYPES:
            value = g.split(":")[name_index]
            if len(inside) > 0:
                value = get_gt_sum(value, inside)
            ws.cell(column=col, row=row, value=value)
            is_single_value.append(col)
            col+=1
            #outline.append(value)
        #print(*[s.data.RC if s.data.RC is not None else "." for s in variant.samples], sep="\t", end="")
    
    for i in is_single_value:
        value = ws.cell(row=row,column=i).value
        for j in range(row + 1, next_row):
            ws.cell(row=j,column=i, value=value)
        #ws.merge_cells(start_row=row, start_column=i, end_row=next_row - 1, end_column=i)

    #ws.column_dimensions['F'].auto_size = True
    #ws.column_dimensions['F'].bestFit = True
    row = next_row
    n_id += 1
wb.save(filename = args.out)